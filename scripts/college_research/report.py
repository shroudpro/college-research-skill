from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


SUMMARY_HEADERS = [
    "大学",
    "大学层次",
    "地理位置",
    "分校区情况",
    "校区具体地址",
    "生活条件",
    "有无空调",
    "宿舍几人间",
    "有无独立卫浴",
    "有无早晚自习",
    "有无早操",
    "宵禁情况",
    "到点是否断电断网",
    "信息来源",
    "CollegesChat页面",
]

RAW_HEADERS = ["大学", "字段", "问题", "时间", "摘录", "CollegesChat页面"]

MAPPING_ROWS = [
    ["校区具体地址", "学校官网/公开校区资料", "替换原“宿舍地理位置”列；未检索到精确门牌号时明确标注或留空。"],
    ["生活条件", "允许点外卖、交通便利、学校超市、自由补充", "先汇总，再列问题、回答和时间。"],
    ["有无空调", "教室和宿舍有没有空调？", "去掉答卷 ID，只保留时间。"],
    ["宿舍几人间", "宿舍是上床下桌吗？", "优先提取人数、上床下桌等关键词，再列近年摘录。"],
    ["有无独立卫浴", "有独立卫浴吗？没有独立浴室的话，澡堂离宿舍多远？", "优先汇总独卫和澡堂距离，再列近年摘录。"],
    ["有无早晚自习", "有早自习、晚自习吗？", "优先汇总统一要求和学院差异。"],
    ["有无早操", "有晨跑吗？每学期跑步打卡的要求是多少公里，可以骑车吗？", "优先汇总晨跑和跑步打卡要求。"],
    ["宵禁情况", "现阶段学校的门禁情况如何？宿舍晚上查寝吗，封寝吗，晚归能回去吗？", "优先汇总门禁、查寝、晚归。"],
    ["到点是否断电断网", "每天断电断网吗，几点开始断？宿舍限电情况？", "优先汇总断电断网和限电情况。"],
]

WIDTHS = {
    "大学": 16,
    "大学层次": 34,
    "地理位置": 20,
    "分校区情况": 26,
    "校区具体地址": 48,
    "生活条件": 70,
    "有无空调": 62,
    "宿舍几人间": 62,
    "有无独立卫浴": 70,
    "有无早晚自习": 62,
    "有无早操": 62,
    "宵禁情况": 70,
    "到点是否断电断网": 68,
    "信息来源": 62,
    "CollegesChat页面": 36,
    "字段": 18,
    "问题": 45,
    "时间": 12,
    "摘录": 78,
    "Excel字段": 18,
    "来源问题或资料": 56,
    "处理规则": 76,
}


def write_rows(ws, headers: list[str], rows: list[dict[str, Any] | list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        if isinstance(row, dict):
            ws.append([row.get(header, "") for header in headers])
        else:
            ws.append(row)


def apply_sheet_style(ws, headers: list[str], table_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        max_lines = 1
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            width = WIDTHS.get(headers[cell.column - 1], 24)
            max_lines = max(max_lines, max(1, len(str(cell.value or "")) // max(12, int(width * 1.3)) + 1))
        ws.row_dimensions[row[0].row].height = min(360, max(42, max_lines * 18))
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = WIDTHS.get(header, 24)
    ws.freeze_panes = "B2" if ws.title in {"汇总表", "原始摘录"} else "A2"
    if ws.max_row >= 2:
        table_ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)


def build_workbook(dataset: dict[str, Any], output_path: Path) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    summary = workbook.create_sheet("汇总表")
    write_rows(summary, SUMMARY_HEADERS, dataset.get("summary_rows", []))
    apply_sheet_style(summary, SUMMARY_HEADERS, "SummaryTable")

    raw = workbook.create_sheet("原始摘录")
    write_rows(raw, RAW_HEADERS, dataset.get("raw_rows", []))
    apply_sheet_style(raw, RAW_HEADERS, "RawEvidenceTable")

    mapping = workbook.create_sheet("字段映射")
    write_rows(mapping, ["Excel字段", "来源问题或资料", "处理规则"], MAPPING_ROWS)
    apply_sheet_style(mapping, ["Excel字段", "来源问题或资料", "处理规则"], "FieldMappingTable")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def validate_workbook(path: Path, expected_schools: int | None = None) -> list[str]:
    errors: list[str] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    if "说明" in sheet_names:
        errors.append("不应存在“说明”sheet")
    for required in ["汇总表", "原始摘录", "字段映射"]:
        if required not in sheet_names:
            errors.append(f"缺少 sheet：{required}")
    if "汇总表" in sheet_names:
        ws = workbook["汇总表"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        forbidden_headers = {"CollegesChat命中状态", "问卷证据概况", "冲突提示"}
        overlap = forbidden_headers.intersection(set(headers))
        if overlap:
            errors.append(f"存在应删除的列：{sorted(overlap)}")
        data_rows = max(0, ws.max_row - 1)
        if expected_schools is not None and data_rows != expected_schools:
            errors.append(f"学校数量不一致：期望 {expected_schools}，实际 {data_rows}")
    id_pattern = re.compile(r"A\d{4,6}")
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = str(cell.value or "")
                if id_pattern.search(value):
                    errors.append(f"发现答卷 ID 残留：{ws.title}!{cell.coordinate}")
                    return errors
    return errors
