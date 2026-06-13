from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook


SCHOOL_SUFFIXES = ("大学", "学院", "学校", "研究院")
HEADER_HINTS = ("大学", "院校", "学校", "院校名称", "学校名称")
NOISE_WORDS = ("专业", "代码", "计划", "备注", "层次", "地址", "官网", "来源")


def normalize_school(value: object) -> str:
    text = re.sub(r"\s+", "", str(value or ""))
    text = text.strip("：:，,。；;")
    return text


def looks_like_school(value: object) -> bool:
    text = normalize_school(value)
    if len(text) < 3 or len(text) > 30:
        return False
    if any(word in text for word in NOISE_WORDS):
        return False
    return text.endswith(SCHOOL_SUFFIXES) or "（原" in text


def find_school_column(rows: list[list[object]]) -> int | None:
    for row in rows[:12]:
        for index, value in enumerate(row):
            text = normalize_school(value)
            if any(hint == text or hint in text for hint in HEADER_HINTS):
                return index
    return None


def extract_schools_from_xlsx(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    schools: list[str] = []
    seen: set[str] = set()

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        column_index = find_school_column([list(row) for row in rows])
        candidates: list[object] = []
        if column_index is not None:
            candidates = [row[column_index] for row in rows if len(row) > column_index]
        else:
            for row in rows:
                candidates.extend(row)
        for value in candidates:
            school = normalize_school(value)
            if looks_like_school(school) and school not in seen:
                schools.append(school)
                seen.add(school)
    return schools
